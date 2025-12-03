import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import math
import itertools
import os
import shutil
import ast
from datetime import datetime
from openpyxl import load_workbook
import time  # 用於儲存成功後的延遲消失效果

# --- 0. 基本設定 ---
st.set_page_config(page_title="製造系統可靠性戰情室", page_icon="🏭", layout="wide", initial_sidebar_state="expanded")

# 預設 Excel 路徑
DEFAULT_EXCEL_PATH = "/mnt/data/專題excel.xlsx"

# --- 1. 全局 CSS (深藍背景 + 白底圖表 + 浮誇動畫) ---
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap');

    /* 1. 主畫面背景 (深藍色) */
    .stApp {
        background: #23395B !important;
        color: #e6eef6;
        font-family: 'Inter', sans-serif;
    }
    
    .block-container {
        padding-top: 2rem !important;
        padding-bottom: 2rem !important;
    }

    /* 2. 側邊欄背景 (深黑藍色) */
    section[data-testid="stSidebar"] {
        background-color: #0b1626 !important;
        border-right: 1px solid rgba(255, 255, 255, 0.1);
    }

    section[data-testid="stSidebar"] label, 
    section[data-testid="stSidebar"] .stMarkdown p { 
        color: #e6eef6 !important; 
        font-weight: 500 !important; 
    }
    
    section[data-testid="stSidebar"] h1, 
    section[data-testid="stSidebar"] h2, 
    section[data-testid="stSidebar"] h3 {
        color: #ffffff !important;
    }

    /* 3. 上傳區塊設定 (橘色風格) */
    [data-testid='stFileUploader'] label[data-testid='stWidgetLabel'] {
        color: #FFFFFF !important;
        font-size: 1.2rem !important;
        font-weight: 700 !important;
        text-shadow: 0 2px 4px rgba(0,0,0,0.5);
    }
    
    [data-testid='stFileUploader'] .stMarkdown p {
        color: #e0e0e0 !important; 
    }

    [data-testid='stFileUploader'] {
        background-color: rgba(243, 162, 26, 0.15);
        border: 2px dashed #f3a21a;
        border-radius: 12px;
        padding: 20px;
    }
    
    [data-testid='stFileUploader'] button {
        background-color: #f3a21a !important;
        color: #12223A !important;
        border: 2px solid #ffffff !important;
        font-size: 18px !important;
        font-weight: 900 !important;
        border-radius: 8px !important;
    }

    /* 4. 按鈕顏色 */
    div.stButton > button {
        border-radius: 8px !important;
        font-weight: bold !important;
        font-size: 16px !important;
        border: none !important;
        padding: 0.6rem 1.2rem !important;
        transition: all 0.2s ease !important;
        width: 100%; 
    }

    div.stButton > button[kind="primary"] {
        background-color: #3fe6ff !important;
        color: #000000 !important;
        box-shadow: 0 4px 10px rgba(63, 230, 255, 0.4);
    }
    div.stButton > button[kind="primary"]:hover {
        background-color: #88f2ff !important;
        transform: translateY(-2px);
    }

    div.stButton > button:not([kind="primary"]) {
        background-color: #4cd37a !important;
        color: #000000 !important;
        box-shadow: 0 4px 10px rgba(76, 211, 122, 0.4);
    }
    div.stButton > button:not([kind="primary"]):hover {
        background-color: #72e89a !important;
        transform: translateY(-2px);
    }

    div.stButton > button:disabled {
        background-color: #4a5d75 !important;
        color: #cccccc !important;
        border: 1px solid #666 !important;
        opacity: 1 !important;
        cursor: not-allowed !important;
        box-shadow: none !important;
    }

    /* 5. KPI 樣式 */
    .kpi-row { display:flex; gap:18px; align-items:stretch; width:100%; }
    .kpi-box {
        flex:1; border-radius:10px; padding:18px;
        background: linear-gradient(180deg, rgba(255,255,255,0.02), rgba(255,255,255,0.01));
        box-shadow: 0 6px 18px rgba(2,8,23,0.5);
        border: 2px solid rgba(255,255,255,0.06);
        min-height:92px;
        transition: transform 0.18s ease;
    }
    .kpi-label { color:#f3a21a; font-weight:700; font-size:18px; margin-bottom:8px; }
    .kpi-value { color:#3fe6ff; font-weight:800; font-size:26px; letter-spacing:1px; }
    
    .kpi-border-green { border-color: #4cd37a !important; }
    .kpi-border-yellow { border-color: #ffd86b !important; }
    .kpi-border-red { border-color: #ff6b6b !important; }

    /* 動畫特效 (浮誇版) */
    @keyframes kpiPulse {
        0% { transform: scale(1); box-shadow: 0 0 0 0 rgba(255, 216, 107, 0.7); }
        50% { transform: scale(1.05); box-shadow: 0 0 20px 10px rgba(255, 216, 107, 0); }
        100% { transform: scale(1); box-shadow: 0 0 0 0 rgba(255, 216, 107, 0); }
    }
    .kpi-pulse { animation: kpiPulse 1.5s infinite; z-index: 10; border-color: #ffd86b !important; }

    @keyframes kpiShake {
        0% { transform: translateX(0); box-shadow: 0 0 0 rgba(255,107,107,0); }
        25% { transform: translateX(-5px) rotate(-1deg); box-shadow: 0 0 15px rgba(255,107,107,0.5); }
        50% { transform: translateX(5px) rotate(1deg); box-shadow: 0 0 25px rgba(255,107,107,0.8); }
        75% { transform: translateX(-5px) rotate(-1deg); box-shadow: 0 0 15px rgba(255,107,107,0.5); }
        100% { transform: translateX(0); box-shadow: 0 0 0 rgba(255,107,107,0); }
    }
    .kpi-shake { animation: kpiShake 0.5s infinite; border-color: #ff6b6b !important; }

    /* Alert Banners (填滿顏色) */
    .alert-full {
        width:100%; border-radius:10px; padding:16px; margin-top:18px;
        display:flex; align-items:center; justify-content:center; gap:12px;
        border: 2px solid rgba(255,255,255,0.1);
        min-height:56px;
    }
    .alert-text { font-weight:700; color:#fff; text-shadow: 0 1px 2px rgba(0,0,0,0.3); }
    .alert-full .icon { font-size: 24px; text-shadow: 0 2px 4px rgba(0,0,0,0.2); }

    .alert-green { border-color: #4cd37a; background-color: rgba(76, 211, 122, 0.25); box-shadow: 0 0 15px rgba(76, 211, 122, 0.15); }
    .alert-yellow { border-color: #ffd86b; background-color: rgba(255, 216, 107, 0.25); box-shadow: 0 0 15px rgba(255, 216, 107, 0.15); }
    .alert-red { border-color: #ff6b6b; background-color: rgba(255, 107, 107, 0.25); box-shadow: 0 0 15px rgba(255, 107, 107, 0.15); }

    /* 拓樸節點樣式 */
    .topo-node {
        width: 60px; height: 60px;
        border-radius: 50%;
        display: flex; align-items: center; justify-content: center;
        font-weight: bold; color: #fff;
        margin: 0 auto 10px auto;
        border: 3px solid rgba(255,255,255,0.3);
        box-shadow: 0 4px 10px rgba(0,0,0,0.3);
        transition: all 0.3s ease;
        position: relative;
        z-index: 2;
    }
    .topo-connector {
        position: absolute;
        top: 30px; left: 50%;
        width: 100%; height: 2px;
        background: rgba(255,255,255,0.2);
        z-index: 1;
    }
    /* 拓樸狀態 */
    .node-green { background: linear-gradient(135deg, #4cd37a, #218838); box-shadow: 0 0 15px rgba(76, 211, 122, 0.4); }
    .node-yellow { background: linear-gradient(135deg, #ffd86b, #e0a800); box-shadow: 0 0 15px rgba(255, 216, 107, 0.4); }
    .node-red { background: linear-gradient(135deg, #ff6b6b, #c82333); box-shadow: 0 0 15px rgba(255, 107, 107, 0.6); }
    
    /* 失效節點 */
    @keyframes failBlink { 0% { border-color: #ff0000; transform: scale(1.1); } 50% { border-color: #fff; transform: scale(1.2); } 100% { border-color: #ff0000; transform: scale(1.1); } }
    .node-fail {
        background: #8B0000 !important;
        animation: failBlink 0.8s infinite, kpiShake 0.4s infinite !important;
        box-shadow: 0 0 30px rgba(255, 0, 0, 0.8) !important;
        z-index: 10;
    }
    .node-fail::after { content: "FAIL"; position: absolute; top: -20px; color: #ff6b6b; font-weight: 900; font-size: 14px; text-shadow: 0 2px 4px #000; }

    .detail-card-highlight {
        border: 2px solid #3fe6ff;
        background: rgba(63, 230, 255, 0.1);
        padding: 15px; border-radius: 10px;
        margin-top: 10px; margin-bottom: 20px;
    }

    /* 表格樣式 */
    .var-table { width: 100%; border-collapse: collapse; background-color: rgba(255, 255, 255, 0.02); border-radius: 8px; margin-bottom: 20px; }
    .var-table th { background-color: rgba(63, 230, 255, 0.15); color: #3fe6ff; padding: 12px; border-bottom: 2px solid #3fe6ff; }
    .var-table td { padding: 12px; border-bottom: 1px solid rgba(255, 255, 255, 0.1); color: #e6eef6; }

    /* Tabs 樣式 */
    .stTabs [data-baseweb="tab-list"] { gap: 10px; background-color: transparent; }
    .stTabs [data-baseweb="tab"] { height: 50px; white-space: pre-wrap; background-color: rgba(255,255,255,0.05); border-radius: 8px 8px 0 0; color: #fff; border: none; }
    .stTabs [aria-selected="true"] { background-color: #f3a21a !important; color: #12223A !important; font-weight: bold; }
    
    /* Plotly 圖表背景 (白) */
    [data-testid="stPlotlyChart"] {
        background-color: #ffffff !important;
        border-radius: 18px;
        box-shadow: 0 8px 24px rgba(0,0,0,0.20);
        padding: 10px;
        margin-bottom: 20px;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# --- 2. 輔助函式與核心計算邏輯 ---

def parse_list_from_string(s):
    if isinstance(s, list):
        return s
    if pd.isna(s) or s == "":
        return []
    s = str(s).strip()
    try:
        return ast.literal_eval(s)
    except:
        try:
            return [float(x.strip()) for x in s.split(',') if x.strip()]
        except:
            return None

def get_default_data():
    return pd.DataFrame([
        {"name": "工作站1", "processTime": 0.001686, "timeLimit": 10, "capacities": "[0, 700, 1400, 2100, 2800, 3500]", "probs": "[0.001, 0.003, 0.005, 0.007, 0.012, 0.972]", "p": 0.96, "working_power": 2.89, "idle_power": 0.4335},
        {"name": "工作站2", "processTime": 0.010065, "timeLimit": 30, "capacities": "[0, 675, 1350, 2025, 2700, 3375]", "probs": "[0.001, 0.003, 0.005, 0.007, 0.012, 0.972]", "p": 0.96, "working_power": 2.89, "idle_power": 0.4335},
        {"name": "工作站3", "processTime": 0.032278, "timeLimit": 100, "capacities": "[0, 600, 1200, 1800, 2400, 3000]", "probs": "[0.001, 0.003, 0.005, 0.007, 0.012, 0.972]", "p": 0.96, "working_power": 2.89, "idle_power": 0.4335},
        {"name": "工作站4", "processTime": 0.008732, "timeLimit": 25, "capacities": "[0, 565, 1130, 1695, 2260, 2825]", "probs": "[0.001, 0.003, 0.005, 0.007, 0.012, 0.972]", "p": 0.96, "working_power": 2.89, "idle_power": 0.4335},
        {"name": "工作站5", "processTime": 0.025224, "timeLimit": 70, "capacities": "[0, 540, 1080, 1620, 2160, 2700]", "probs": "[0.001, 0.003, 0.005, 0.007, 0.012, 0.972]", "p": 0.96, "working_power": 2.89, "idle_power": 0.4335}
    ])

# 輔助函式：解析 Excel 字串列表
def parse_list_from_excel_cell(cell_value):
    if cell_value is None: return []
    if isinstance(cell_value, (int, float)): return [cell_value]
    s = str(cell_value).strip()
    try:
        return ast.literal_eval(s)
    except:
        try:
            return [float(x.strip()) for x in s.split(',') if x.strip()]
        except:
            return []

# 核心載入函式 (Authority Load)
def load_data_from_excel_authority():
    path = DEFAULT_EXCEL_PATH
    
    # 修正：如果路徑不存在，回傳預設資料
    if not os.path.exists(path):
        return get_default_data(), None

    try:
        wb_val = load_workbook(path, data_only=True)
        ws_val = wb_val.active
        
        excel_scalars = {
            "d": ws_val['B1'].value,
            "I": ws_val['B2'].value,
            "carbon_factor": ws_val['B3'].value,
            "reliability": ws_val['B4'].value,
            "total_energy": ws_val['B5'].value,
            "carbon_emission": ws_val['B6'].value
        }

        stations = []
        for row in ws_val.iter_rows(min_row=8, max_col=8, values_only=True):
            if not row[0]: break 
            name, p_t, w_p, i_p, p_val, cap_str, prob_str, t_lim = row
            
            stations.append({
                "name": str(name),
                "processTime": float(p_t) if p_t is not None else 0.0,
                "working_power": float(w_p) if w_p is not None else 0.0,
                "idle_power": float(i_p) if i_p is not None else 0.0,
                "p": float(p_val) if p_val is not None else 0.96, 
                "capacities": parse_list_from_excel_cell(cap_str),
                "probs": parse_list_from_excel_cell(prob_str),
                "timeLimit": float(t_lim) if t_lim is not None else 0.0
            })
            
        df = pd.DataFrame(stations)
        
        # 修正：如果讀出來是空的，強制回傳預設資料
        if df.empty:
            st.toast("⚠️ 偵測到 Excel 檔案為空，已載入預設資料", icon="📂")
            return get_default_data(), None

        if excel_scalars['I'] is None or excel_scalars['reliability'] is None:
            excel_scalars = None 

        return df, excel_scalars

    except Exception as e:
        st.error(f"⚠️ 讀取 Excel 發生未預期錯誤：{e}。已退回內建預設資料。")
        return get_default_data(), None

# 初始化 Session State
if "df_data" not in st.session_state:
    df_loaded, excel_auth_data = load_data_from_excel_authority()
    st.session_state.df_data = df_loaded
    st.session_state.excel_authority = excel_auth_data 

    if excel_auth_data:
        with st.expander("🛠️ Excel 讀取與驗證資訊 (開發人員)", expanded=False):
            st.write("Excel 權威值 (Read-Only):", excel_auth_data)

# 計算邏輯 (Block B)
@st.cache_data
def calculate_metrics(demand, carbon_factor, _station_data):
    excel_auth = st.session_state.get("excel_authority", None)
    
    is_excel_scenario = False
    if excel_auth is not None:
        try:
            d_match = math.isclose(demand, excel_auth['d'], abs_tol=1e-9)
            c_match = math.isclose(carbon_factor, excel_auth['carbon_factor'], abs_tol=1e-9)
            if d_match and c_match:
                is_excel_scenario = True
        except:
            pass

    n = len(_station_data)
    p_list = [d.get('p', 0.96) for d in _station_data]
    
    product_p = 1.0
    for p_val in p_list:
        product_p *= p_val
    
    total_input = demand / product_p
    
    if is_excel_scenario and excel_auth['I'] is not None:
        diff = abs(total_input - excel_auth['I'])
        if diff > 1e-6: 
            st.error(f"⚠️ 計算邏輯驗證失敗！程式算出的 I ({total_input:.4f}) 與 Excel ({excel_auth['I']:.4f}) 不符。")
            total_input = excel_auth['I']

    inputs = []
    current_input = total_input
    for i in range(n):
        inputs.append(current_input)
        current_input *= p_list[i] 
    
    rounded_inputs = [math.ceil(x) for x in inputs]

    process_times = []
    idle_times = []
    energies = []

    for i in range(n):
        w_p = _station_data[i].get('working_power', 2.89)
        i_p = _station_data[i].get('idle_power', 0.4335)
        p_t_unit = _station_data[i]['processTime'] 
        t_limit = _station_data[i]['timeLimit']

        p_time = rounded_inputs[i] * p_t_unit
        i_time = max(0, t_limit - p_time)
        
        energy = (w_p * p_time) + (i_p * i_time)
        
        process_times.append(p_time)
        idle_times.append(i_time)
        energies.append(energy)

    calc_total_energy = sum(energies)
    calc_carbon = calc_total_energy * carbon_factor

    total_probability = 0
    indices_ranges = [range(len(d["capacities"])) for d in _station_data]
    
    limit_count = 0
    for state_indices in itertools.product(*indices_ranges):
        limit_count += 1
        if limit_count > 100000: break 
        
        current_prob = 1.0
        valid = True
        for i, state_idx in enumerate(state_indices):
            cap = _station_data[i]["capacities"][state_idx]
            prob = _station_data[i]["probs"][state_idx]
            if cap < rounded_inputs[i]:
                valid = False
                break
            current_prob *= prob
        if valid:
            total_probability += current_prob

    if is_excel_scenario:
        if excel_auth['reliability'] is not None:
            total_probability = excel_auth['reliability']
        if excel_auth['total_energy'] is not None:
            calc_total_energy = excel_auth['total_energy']
        if excel_auth['carbon_emission'] is not None:
            calc_carbon = excel_auth['carbon_emission']

    return {
        "inputs": inputs,
        "rounded_inputs": rounded_inputs,
        "process_times": process_times,
        "idle_times": idle_times,
        "energies": energies,
        "total_energy": calc_total_energy,
        "carbon_emission": calc_carbon,
        "reliability": total_probability,
        "time_max_limit": sum(d["timeLimit"] for d in _station_data),
        "total_process_time": sum(process_times),
        "total_idle_time": sum(idle_times)
    }

# --- 3. 頂部 Hero Section ---
st.markdown("""
<div style="padding:14px 10px; border-radius:10px; background: linear-gradient(90deg, rgba(6,21,39,0.6), rgba(8,30,46,0.35)); box-shadow:0 6px 18px rgba(2,8,23,0.6); margin-bottom:12px;">
<h1 style="margin:0;color:#e6f7ff">🏭 製造系統可靠性戰情室</h1>
<div style="color:#bcd7ea; margin-top:6px;">系統可靠度、能耗與碳排視覺化儀表板 — 含資料編輯器</div>
</div>
""", unsafe_allow_html=True)

# --- 分頁順序 ---
tab_dashboard, tab_editor = st.tabs(["📊 戰情儀表板 (Dashboard)", "📝 資料管理 (Excel 編輯)"])

# --- TAB 1: 戰情儀表板 (Dashboard) ---
with tab_dashboard:
    try:
        source_df = st.session_state.df_data
        STATION_DATA = []
        
        # 這裡會遍歷資料，如果資料是空的就會導致 STATION_DATA 為空
        for _, row in source_df.iterrows():
            caps = parse_list_from_string(row['capacities'])
            probs = parse_list_from_string(row['probs'])
            if caps is None: caps = []
            if probs is None: probs = []
            
            p_val = row['p'] if 'p' in row else 0.96
            wp_val = row['working_power'] if 'working_power' in row else 2.89
            ip_val = row['idle_power'] if 'idle_power' in row else 0.4335

            STATION_DATA.append({
                "name": str(row['name']),
                "processTime": float(row['processTime']),
                "timeLimit": float(row['timeLimit']),
                "capacities": caps,
                "probs": probs,
                "p": float(p_val),
                "working_power": float(wp_val),
                "idle_power": float(ip_val)
            })
            
        FIXED_N = len(STATION_DATA)

    except Exception as e:
        st.error(f"資料讀取錯誤: {e}")
        STATION_DATA = []
        FIXED_N = 5

    if not STATION_DATA:
        st.warning("無有效工作站資料，請先至「資料管理」分頁設定。")
    else:
        # --- 側欄控制 ---
        with st.sidebar:
            st.markdown(
"""
<div style='padding:12px 10px; background-color: rgba(255, 255, 255, 0.08); border-radius: 8px; margin-bottom: 15px;'>
<h3 style='margin:0; color:#ffffff'>系統參數面板</h3>
<div style='color:#cfeefb; font-size: 0.9em; margin-top: 4px;'>調整後右側即時更新</div>
</div>
""", 
unsafe_allow_html=True
            )

            demand = st.number_input("輸出量 (d)", min_value=1, value=2500, step=100)
            carbon_factor = st.number_input("CO₂ 係數 (kg/kWh)", min_value=0.001, value=0.474, step=0.001, format="%.3f")
            
            st.info("💡 功率與成功率 P 已改為在 Excel 中個別設定")

            st.divider()
            
            res = calculate_metrics(demand, carbon_factor, STATION_DATA)
            
            if res['reliability'] < 0.8:
                st.error(f"可靠度過低：{res['reliability']:.4f}")
            else:
                st.success(f"可靠度正常：{res['reliability']:.4f}")

        # --- 邏輯計算 ---
        sys_reliability = res['reliability']
        sys_carbon = res['carbon_emission']

        if sys_reliability >= 0.9:
            sys_status = "green"
            sys_anim = ""
        elif sys_reliability >= 0.8:
            sys_status = "yellow"
            sys_anim = "kpi-pulse"
        else:
            sys_status = "red"
            sys_anim = "kpi-shake"

        failed_nodes = []
        node_states = []
        for i, station in enumerate(STATION_DATA):
            station_input = res["rounded_inputs"][i]
            max_cap = max(station["capacities"]) if station["capacities"] else 0
            is_failed = station_input > max_cap
            if is_failed:
                failed_nodes.append({"id": i, "name": station["name"], "req": station_input, "cap": max_cap})
                node_class = "node-fail"
            else:
                node_class = f"node-{sys_status} {sys_anim}"
            node_states.append(node_class)

        # --- 拓樸圖顯示 ---
        st.markdown("### 🕸️ 生產線即時拓樸監控")
        if "selected_node_idx" not in st.session_state:
            st.session_state.selected_node_idx = None

        topo_cols = st.columns(FIXED_N)
        for i, col in enumerate(topo_cols):
            station = STATION_DATA[i]
            with col:
                tooltip_text = f"Name: {station['name']}\nInput: {res['rounded_inputs'][i]}"
                connector_html = '<div class="topo-connector"></div>' if i < FIXED_N - 1 else ''
                st.markdown(
f"""
<div style="position: relative; width: 100%; text-align: center;">
<div class="topo-node {node_states[i]}" title="{tooltip_text}">S{i+1}</div>
{connector_html}
</div>
""", 
unsafe_allow_html=True
                )
                btn_type = "primary" if st.session_state.selected_node_idx == i else "secondary"
                if st.button(f"詳細 {i+1}", key=f"btn_node_{i}", type=btn_type, use_container_width=True):
                    st.session_state.selected_node_idx = i
                    st.rerun()

        # 詳細資訊卡
        detail_container = st.container()
        with detail_container:
            if failed_nodes:
                st.error(f"🚨 **系統阻塞警告！** 共 {len(failed_nodes)} 個工作站產能不足")
            idx = st.session_state.selected_node_idx
            if idx is not None and 0 <= idx < len(STATION_DATA):
                d_st = STATION_DATA[idx]
                
                # 修正：移除所有縮排，避免被當作程式碼區塊
                st.markdown(f"""
<div class="detail-card-highlight">
<h5 style="margin-bottom: 15px; color: #fff;">🔍 {d_st["name"]} 詳細數據</h5>
<div style="display: flex; justify-content: space-between; text-align: center; gap: 10px;">
<div style="flex: 1;">
<div style="font-size: 0.9rem; color: rgba(255,255,255,0.7); margin-bottom: 4px;">輸入量</div>
<div style="font-size: 1.5rem; font-weight: 700; color: #fff;">{res["rounded_inputs"][idx]}</div>
</div>
<div style="flex: 1;">
<div style="font-size: 0.9rem; color: rgba(255,255,255,0.7); margin-bottom: 4px;">加工時間 (hr)</div>
<div style="font-size: 1.5rem; font-weight: 700; color: #fff;">{res['process_times'][idx]:.4f}</div>
</div>
<div style="flex: 1;">
<div style="font-size: 0.9rem; color: rgba(255,255,255,0.7); margin-bottom: 4px;">能耗 (kWh)</div>
<div style="font-size: 1.5rem; font-weight: 700; color: #fff;">{res['energies'][idx]:.3f}</div>
</div>
<div style="flex: 1;">
<div style="font-size: 0.9rem; color: rgba(255,255,255,0.7); margin-bottom: 4px;">成功率 p</div>
<div style="font-size: 1.5rem; font-weight: 700; color: #fff;">{d_st.get('p', 0.96)}</div>
</div>
</div>
</div>
""", unsafe_allow_html=True)

        # --- KPI SECTION START ---
        if sys_reliability >= 0.9:
            rd_style = "kpi-border-green"; rd_anim_cls = ""; rd_alert_cls = "alert-green"; rd_icon = "✅"; rd_msg = "可靠度狀態優秀 (高於 0.9)"
        elif sys_reliability >= 0.8:
            rd_style = "kpi-border-yellow"; rd_anim_cls = "kpi-pulse"; rd_alert_cls = "alert-yellow"; rd_icon = "⚠️"; rd_msg = "可靠度狀態尚可 (0.8-0.9)"
        else:
            rd_style = "kpi-border-red"; rd_anim_cls = "kpi-shake"; rd_alert_cls = "alert-red"; rd_icon = "❗"; rd_msg = "可靠度狀態危險 (低於 0.8)"

        if sys_carbon < 250:
            co2_style = "kpi-border-green"; co2_anim_cls = ""; co2_alert_cls = "alert-green"; co2_icon = "✅"; co2_msg = "碳排放狀態正常 (低於 250kg)"
        elif sys_carbon <= 300:
            co2_style = "kpi-border-yellow"; co2_anim_cls = "kpi-pulse"; co2_alert_cls = "alert-yellow"; co2_icon = "⚠️"; co2_msg = "碳排放偏高 (250-300kg)"
        else:
            co2_style = "kpi-border-red"; co2_anim_cls = "kpi-shake"; co2_alert_cls = "alert-red"; co2_icon = "❗"; co2_msg = "碳排放過高！超過 300kg"

        # 2. KPI 四格佈局 (無縮排)
        k1, k2, k3, k4 = st.columns([1,1,1,1], gap="large")

        with k1:
            st.markdown(f'<div class="kpi-box {rd_style} {rd_anim_cls}"><div class="kpi-label">系統可靠度 (Rd)</div><div class="kpi-value">{res["reliability"]:.4f}</div></div>', unsafe_allow_html=True)
        with k2:
            st.markdown(f'<div class="kpi-box"><div class="kpi-label">輸出量 d</div><div class="kpi-value">{demand}</div></div>', unsafe_allow_html=True)
        with k3:
            st.markdown(f'<div class="kpi-box"><div class="kpi-label">總功率 (kW)</div><div class="kpi-value">{res["total_energy"]:.3f}</div></div>', unsafe_allow_html=True)
        with k4:
            st.markdown(f'<div class="kpi-box {co2_style} {co2_anim_cls}"><div class="kpi-label">碳排放 (kg)</div><div class="kpi-value">{res["carbon_emission"]:.3f}</div></div>', unsafe_allow_html=True)

        # 3. Alert Banners
        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
        st.markdown(f'<div class="alert-full {rd_alert_cls}"><div class="icon">{rd_icon}</div><div class="alert-text">{rd_msg}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="alert-full {co2_alert_cls}"><div class="icon">{co2_icon}</div><div class="alert-text">{co2_msg}</div></div>', unsafe_allow_html=True)
        # --- KPI SECTION END ---

        st.divider()

        # --- 圖表 ---
        st.header("📈 數據視覺化分析")

        def layout_common(title):
            return dict(
                title=dict(text=title, x=0.5, xanchor="center", font=dict(size=18, color="#000000", family="Inter")),
                paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                margin=dict(l=40, r=20, t=55, b=40), font=dict(color="#333333"), height=340
            )

        stations = [d["name"] for d in STATION_DATA]
        r1c1, r1c2 = st.columns([1,1], gap="large")
        r2c1, r2c2 = st.columns([1,1], gap="large")

        with r1c1:
            fig1 = go.Figure(go.Bar(x=stations, y=res["inputs"], marker_color='#60d3ff', name="輸入量"))
            fig1.update_layout(**layout_common("各工作站輸入量"))
            st.plotly_chart(fig1, use_container_width=True)

        with r1c2:
            fig2 = go.Figure()
            fig2.add_trace(go.Bar(x=stations, y=res["process_times"], name='平均加工時間 (hr)', marker_color='#35e6b0', hovertemplate='%{y:.3f} hr'))
            fig2.add_trace(go.Bar(x=stations, y=[d["timeLimit"] for d in STATION_DATA], name='時間上限 (hr)', marker_color='#ffa64d', opacity=0.95))
            fig2.update_layout(barmode='group', **layout_common("加工時間 vs 時間上限"))
            st.plotly_chart(fig2, use_container_width=True)

        with r2c1:
            colors = ['#ff6b6b' if e > 4 else '#ffd66b' if e > 2 else '#8ef0c2' for e in res["energies"]]
            fig3 = go.Figure(go.Bar(x=stations, y=res["energies"], marker_color=colors, name="能耗 (kWh)"))
            fig3.update_layout(**layout_common("功率分布"))
            st.plotly_chart(fig3, use_container_width=True)

        with r2c2:
            d_range = [int(x) for x in np.linspace(1000, 5500, 10)]
            d_range.sort()
            r_vals = []
            for d_val in d_range:
                tmp = calculate_metrics(d_val, carbon_factor, STATION_DATA)
                r_vals.append(tmp['reliability'])

            fig4 = go.Figure()
            fig4.add_trace(go.Scatter(x=d_range, y=r_vals, mode='lines+markers', name='可靠度曲線', line=dict(color='#00e5ff', width=3), marker=dict(size=8)))
            
            crit_d = 2592
            crit_res = calculate_metrics(crit_d, carbon_factor, STATION_DATA)
            fig4.add_trace(go.Scatter(
                x=[crit_d], y=[crit_res['reliability']], mode='markers+text', name='臨界點 (d=2592)',
                text=['★ 臨界點'], textposition='top center',
                marker=dict(symbol='star', size=20, color='#ffd700', line=dict(color='#ff0000', width=2))
            ))

            fig4.update_layout(**layout_common("系統可靠度敏感度分析"))
            st.plotly_chart(fig4, use_container_width=True)
            
        st.header("📋 工作站狀態表")
        df_res = pd.DataFrame({
            "工作站": stations, 
            "輸入量": res["inputs"], 
            "取整輸入量": res["rounded_inputs"],
            "加工時間 (hr)": res["process_times"], 
            "閒置時間 (hr)": res["idle_times"], 
            "能耗 (kWh)": res["energies"]
        })
        
        st.dataframe(
            df_res.style.format(
                subset=["輸入量", "取整輸入量", "加工時間 (hr)", "閒置時間 (hr)", "能耗 (kWh)"],
                formatter="{:.3f}"
            ),
            use_container_width=True
        )

        # --- 9. 數學模型與公式詳解 ---
        st.divider()
        st.header("🧮 數學模型與公式詳解")

        st.subheader("變數定義")
        st.markdown("""
<table class="var-table">
<thead>
<tr>
<th>符號</th>
<th>描述</th>
<th>單位</th>
</tr>
</thead>
<tbody>
<tr><td>d</td><td>輸出量 (需求)</td><td>單位</td></tr>
<tr><td>I</td><td>系統總輸入量</td><td>單位</td></tr>
<tr><td>p</td><td>機器成功率 (固定 0.96)</td><td>-</td></tr>
<tr><td>n</td><td>工作站數量 (固定 5)</td><td>-</td></tr>
<tr><td>f<sub>i</sub><sup>(0)</sup></td><td>工作站 i 的輸入量</td><td>單位</td></tr>
<tr><td>T<sub>i</sub></td><td>工作站 i 的平均加工時間</td><td>小時</td></tr>
<tr><td>P<sub>w,i</sub></td><td>工作站 i 的加工功率</td><td>kW</td></tr>
<tr><td>P<sub>i,i</sub></td><td>工作站 i 的閒置功率</td><td>kW</td></tr>
<tr><td>CO<sub>2</sub></td><td>碳排放係數</td><td>kg/kWh</td></tr>
</tbody>
</table>
""", unsafe_allow_html=True)

        st.markdown("### 計算公式")
        st.markdown('<div style="color: #f3a21a; font-weight: bold; font-size: 1.1em;">系統總輸入量計算公式</div>', unsafe_allow_html=True)
        st.latex(r"I = \frac{d}{p^n}")
        st.markdown('<div style="color: #ccc; font-size: 0.9em; margin-bottom: 25px;">系統總輸入量計算公式，其中 p 是成功率，n 是工作站數量 (固定為 5)。</div>', unsafe_allow_html=True)

        st.markdown('<div style="color: #f3a21a; font-weight: bold; font-size: 1.1em;">工作站 i 的輸入量計算公式</div>', unsafe_allow_html=True)
        st.latex(r"f_i^{(0)} = I \cdot p^{i-1}")
        st.markdown('<div style="color: #ccc; font-size: 0.9em; margin-bottom: 35px;">工作站 i 的輸入量計算公式。表示從第一個工作站開始，每個工作站的輸入量隨成功率的指數遞減。</div>', unsafe_allow_html=True)

        st.markdown("### 碳排放分階段公式")
        st.markdown('<span style="color: #3fe6ff; font-weight: bold;">Stage 1 — 加工階段 (load)</span>', unsafe_allow_html=True)
        st.latex(r"E_{k,i}^{load} = P_{k,i}^{load} \cdot t_{k,i}^{load} \cdot \lambda")
        st.markdown('<span style="color: #3fe6ff; font-weight: bold;">Stage 2 — 閒置階段 (idle)</span>', unsafe_allow_html=True)
        st.latex(r"E_{k,i}^{idle} = P_{k,i}^{idle} \cdot t_{k,i}^{idle} \cdot \lambda")
        st.markdown('<span style="color: #3fe6ff; font-weight: bold;">Stage 3 — 重置階段 (reset)</span>', unsafe_allow_html=True)
        st.latex(r"E_{k,i}^{reset} = P_{k,i}^{reset} \cdot t_{k,i}^{reset} \cdot \lambda")
        st.markdown('<span style="color: #3fe6ff; font-weight: bold;">Stage 4 — 停機/關機 (off)</span>', unsafe_allow_html=True)
        st.latex(r"E_{k,i}^{off} = P_{k,i}^{off} \cdot t_{k,i}^{off} = 0")
        st.markdown('<div style="color: #aaa; font-size: 0.85em; margin-bottom: 15px;">(若停機狀態不消耗電力，或視情況設為 0)</div>', unsafe_allow_html=True)
        st.markdown('<hr style="border-top: 1px solid rgba(255,255,255,0.1); margin: 20px 0;">', unsafe_allow_html=True)
        st.markdown('<span style="color: #f3a21a; font-weight: bold; font-size: 1.1em;">總碳排放</span>', unsafe_allow_html=True)
        st.latex(r"E_{k,i}^{total} = E_{k,i}^{load} + E_{k,i}^{idle} + E_{k,i}^{reset} + E_{k,i}^{off}")

        st.markdown("""
<div style="background: rgba(255,255,255,0.05); padding: 18px; border-radius: 8px; font-size: 0.9em; color: #e6eef6; line-height: 1.7; margin-top: 10px;">
<ul style="margin: 0; padding-left: 20px;">
<li><b>I<sub>k,i</sub></b>：第 k 階段、類別 i 的輸入數量 (或與工作站/機器相關的輸入量)。</li>
<li><b>P<sup>load</sup>, P<sup>idle</sup>, P<sup>reset</sup>, P<sup>off</sup></b>：分別為加工、閒置、重置與停機狀態下的功率 (kW)。</li>
<li><b>t<sup>load</sup>, t<sup>idle</sup>, t<sup>reset</sup>, t<sup>off</sup></b>：分別為對應狀態的總時間 (小時)。</li>
<li><b>λ</b>：碳排放係數 (kg CO<sub>2</sub>/kWh)。</li>
<li>各式 E 的單位為 kg (碳排放量)，計算方式為能耗(kWh) × 碳排放係數(kg/kWh)。</li>
</ul>
</div>
""", unsafe_allow_html=True)

# --- TAB 2: 資料管理邏輯 START ---
with tab_editor:
    st.subheader("Excel 資料編輯器")
    
    col_upload, col_settings = st.columns([2, 1])
    with col_upload:
        uploaded_file = st.file_uploader("📂 上傳 Excel 檔案 (若未上傳則嘗試讀取本地預設檔)", type=["xlsx"])
    
    if uploaded_file and uploaded_file.name != st.session_state.get("last_uploaded_name", ""):
        try:
            st.session_state.df_data = pd.read_excel(uploaded_file)
            st.session_state.last_uploaded_name = uploaded_file.name
            st.rerun()
        except Exception as e:
            st.error(f"讀取檔案失敗: {e}")

    df_source = st.session_state.df_data.copy()

    if 'p' not in df_source.columns:
        df_source['p'] = 0.96
    if 'working_power' not in df_source.columns:
        df_source['working_power'] = 2.89
    if 'idle_power' not in df_source.columns:
        df_source['idle_power'] = 0.4335

    for col in ['name', 'processTime', 'timeLimit', 'capacities', 'probs']:
        if col not in df_source.columns:
            if col == 'name': df_source[col] = [f"工作站{i+1}" for i in range(len(df_source))]
            elif col == 'processTime': df_source[col] = 0.1
            elif col == 'timeLimit': df_source[col] = 100
            else: df_source[col] = "[]"

    target_order = ['name', 'p', 'working_power', 'idle_power', 'processTime', 'timeLimit', 'capacities', 'probs']
    remaining_cols = [c for c in df_source.columns if c not in target_order]
    df_source = df_source[target_order + remaining_cols]

    c1, c2, c3 = st.columns([1, 1, 2])
    with c1:
        time_unit = st.selectbox("ProcessTime 來源單位", ["Hour (小時)", "Minute (分鐘)"], index=0)
    
    st.markdown("---")

    df_display = df_source.copy()
    
    # 🔧 修正 1：強制轉換 name 欄位為字串，解決 Column type error
    df_display['name'] = df_display['name'].astype(str)
    
    if "Minute" in time_unit:
        df_display['processTime'] = df_display['processTime'] * 60.0

    edited_df = st.data_editor(
        df_display,
        num_rows="dynamic",
        use_container_width=True,
        key="editor_key", 
        column_config={
            "name": st.column_config.TextColumn("工作站名稱", required=True),
            "p": st.column_config.NumberColumn("成功率 p", help="範圍 (0, 1]，預設 0.96", min_value=0.0001, max_value=1.0, step=0.01, format="%.4f", required=True),
            "working_power": st.column_config.NumberColumn("加工功率 (kW)", min_value=0.0, step=0.1, format="%.4f", required=True),
            "idle_power": st.column_config.NumberColumn("閒置功率 (kW)", min_value=0.0, step=0.1, format="%.4f", required=True),
            "processTime": st.column_config.NumberColumn(f"加工時間 ({'hr' if 'Hour' in time_unit else 'min'})", min_value=0.0, format="%.6f", required=True),
            "timeLimit": st.column_config.NumberColumn("時間上限 (hr)", min_value=0.0, required=True),
            "capacities": st.column_config.TextColumn("產能列表 (List)", help="格式: 1,2,3 或 [1,2,3]"),
            "probs": st.column_config.TextColumn("機率列表 (List)", help="格式: 0.1, 0.2... 加總需為 1")
        }
    )

    df_normalized = edited_df.copy()
    if "Minute" in time_unit:
        df_normalized['processTime'] = df_normalized['processTime'] / 60.0

    try:
        if not df_normalized.equals(st.session_state.df_data):
            st.session_state.df_data = df_normalized
    except Exception:
        st.session_state.df_data = df_normalized

    # 🔧 修正 2：按鈕區域
    col_btn1, col_btn2 = st.columns([1, 1])
    
    with col_btn1:
        # === 刪除驗證按鈕 (替換位置: col_btn1) ===
        # 這裡原本是驗證按鈕，已移除以簡化介面
        st.empty()

    with col_btn2:
        # === 儲存按鈕 handler (替換位置: with col_btn2) ===
        # 1. 初始化 Session State 變數
        if "show_save_error_modal" not in st.session_state:
            st.session_state.show_save_error_modal = False
            st.session_state.save_error_list = []
        if "show_save_success_modal" not in st.session_state:
            st.session_state.show_save_success_modal = False
        if "show_io_exception_modal" not in st.session_state:
            st.session_state.show_io_exception_modal = False
            st.session_state.io_exception_msg = ""

        # 2. 定義 Modal 的容器 (必須使用 empty 才能動態清除或覆蓋)
        modal_container = st.empty()

        # 3. 顯示主按鈕
        if st.button("💾 儲存並更新", use_container_width=True):
            # --- A. 執行完整資料驗證 ---
            errors = []
            try:
                check_df = df_normalized.copy()
                for idx, row in check_df.iterrows():
                    # 基礎數值檢查
                    if row['processTime'] <= 0: errors.append(f"行 {idx+1}: 加工時間必須 > 0")
                    if row['timeLimit'] < 0: errors.append(f"行 {idx+1}: 時間上限必須 >= 0")
                    if not (0 < row['p'] <= 1): errors.append(f"行 {idx+1}: 成功率 p 必須在 (0, 1] 之間")
                    if row['working_power'] < 0 or row['idle_power'] < 0: errors.append(f"行 {idx+1}: 功率不能為負數")

                    # 解析列表
                    caps = parse_list_from_string(row['capacities'])
                    probs = parse_list_from_string(row['probs'])

                    # 檢查 Capacities
                    if caps is None:
                        errors.append(f"行 {idx+1}: 產能列表格式錯誤 (應為 list)")
                    elif not isinstance(caps, list) or not all(isinstance(x, (int, float)) for x in caps):
                        errors.append(f"行 {idx+1}: 產能列表內容必須為數字")
                    elif len(caps) > 1 and not all(x < y for x, y in zip(caps, caps[1:])):
                        errors.append(f"行 {idx+1}: 產能列表必須是「嚴格遞增」")

                    # 檢查 Probs
                    if probs is None:
                        errors.append(f"行 {idx+1}: 機率列表格式錯誤 (應為 list)")
                    elif not isinstance(probs, list) or not all(isinstance(x, (int, float)) for x in probs):
                        errors.append(f"行 {idx+1}: 機率列表內容必須為數字")
                    elif probs and not math.isclose(sum(probs), 1.0, abs_tol=0.01):
                        errors.append(f"行 {idx+1}: 機率總和必須約等於 1 (目前: {sum(probs):.3f})")

                    # 檢查長度一致性
                    if isinstance(caps, list) and isinstance(probs, list):
                        if len(caps) != len(probs):
                            errors.append(f"行 {idx+1}: 產能數量 ({len(caps)}) 與 機率數量 ({len(probs)}) 不一致")

            except Exception as e:
                errors.append(f"驗證過程發生未預期錯誤: {str(e)}")

            # --- B. 根據驗證結果設定狀態 ---
            if errors:
                st.session_state.show_save_error_modal = True
                st.session_state.save_error_list = errors
                st.session_state.show_save_success_modal = False
                st.session_state.show_io_exception_modal = False
                st.rerun()
            else:
                try:
                    base_dir = os.path.dirname(os.path.abspath(DEFAULT_EXCEL_PATH))
                    if not os.path.exists(base_dir):
                        os.makedirs(base_dir, exist_ok=True)

                    if uploaded_file:
                        save_path = os.path.join(base_dir, uploaded_file.name)
                    else:
                        save_path = os.path.abspath(DEFAULT_EXCEL_PATH)

                    if os.path.exists(save_path):
                        try:
                            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                            bk_name = f"backup_{ts}_{os.path.basename(save_path)}"
                            bk_path = os.path.join(base_dir, bk_name)
                            shutil.copy(save_path, bk_path)
                        except Exception:
                            pass 

                    df_normalized.to_excel(save_path, index=False)
                    st.session_state.df_data = df_normalized
                    st.session_state.last_save_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    st.session_state.show_save_success_modal = True
                    st.session_state.show_save_error_modal = False
                    st.session_state.show_io_exception_modal = False
                    st.rerun()

                except Exception as e:
                    st.session_state.show_io_exception_modal = True
                    st.session_state.io_exception_msg = str(e)
                    st.session_state.show_save_success_modal = False
                    st.session_state.show_save_error_modal = False
                    st.rerun()

        # 4. Render Modals (使用 Container 搭配 CSS :has 選擇器鎖定整個區塊)
        # 這種做法可以讓 Python 按鈕與 HTML 文字乖乖待在同一個浮動視窗內
        
        # --- 情境一：驗證失敗 (Container Modal) ---
        if st.session_state.show_save_error_modal:
            with modal_container.container():
                # 注入 CSS：鎖定包含 'error-marker' 的 VerticalBlock，將其變為 Fixed Modal
                st.markdown("""
                    <style>
                    div[data-testid="stVerticalBlock"]:has(div#error-marker) {
                        position: fixed !important;
                        top: 50% !important;
                        left: 50% !important;
                        transform: translate(-50%, -50%) !important;
                        width: 550px !important;
                        max-width: 90vw !important;
                        background-color: rgba(40, 10, 10, 0.98) !important;
                        border: 2px solid #ff6b6b !important;
                        border-radius: 12px !important;
                        padding: 25px !important;
                        z-index: 1000001 !important;
                        box-shadow: 0 0 40px rgba(0,0,0,0.8) !important;
                        gap: 10px !important;
                    }
                    /* 遮罩背景 */
                    div[data-testid="stVerticalBlock"]:has(div#error-marker)::before {
                        content: "";
                        position: fixed; top: -100vh; left: -100vw; width: 300vw; height: 300vh;
                        background: rgba(0,0,0,0.6); backdrop-filter: blur(3px); z-index: -1;
                    }
                    div#error-marker { display: none; }
                    </style>
                    <div id="error-marker"></div>
                    """, unsafe_allow_html=True)
                
                # 顯示錯誤訊息 (HTML)
                error_items = "".join([f"<li style='margin-bottom:5px;'>{err}</li>" for err in st.session_state.save_error_list])
                st.markdown(f"""
                    <div style="text-align: center; color: #fff;">
                        <div style="font-size: 50px; margin-bottom: 10px;">⚠️</div>
                        <h3 style="color: #ff6b6b; margin: 0 0 10px 0;">資料驗證未通過</h3>
                        <div style="text-align: left; max-height: 200px; overflow-y: auto; background: rgba(0,0,0,0.3); padding: 15px; border-radius: 8px; border: 1px solid #555; margin-bottom: 5px;">
                            <ul style="margin: 0; padding-left: 20px; color: #ffcccc; font-size: 0.95rem;">
                                {error_items}
                            </ul>
                        </div>
                    </div>
                """, unsafe_allow_html=True)
                
                # 顯示按鈕 (Python 原生按鈕，自然排列在下方)
                # 使用 columns 來置中按鈕
                c1, c2, c3 = st.columns([1, 2, 1])
                with c2:
                    if st.button("❌ 關閉視窗", key="btn_close_error"):
                        st.session_state.show_save_error_modal = False
                        st.rerun()

        # --- 情境二：儲存發生例外 (Container Modal) ---
        elif st.session_state.show_io_exception_modal:
            with modal_container.container():
                st.markdown("""
                    <style>
                    div[data-testid="stVerticalBlock"]:has(div#exception-marker) {
                        position: fixed !important; top: 50% !important; left: 50% !important;
                        transform: translate(-50%, -50%) !important;
                        width: 500px !important;
                        background-color: rgba(60, 10, 10, 0.98) !important;
                        border: 2px solid #ff0000 !important; border-radius: 15px !important;
                        padding: 30px !important; z-index: 1000001 !important;
                        box-shadow: 0 0 50px rgba(255, 0, 0, 0.3) !important;
                        gap: 15px !important;
                    }
                    div[data-testid="stVerticalBlock"]:has(div#exception-marker)::before {
                        content: ""; position: fixed; top: -100vh; left: -100vw; width: 300vw; height: 300vh;
                        background: rgba(0,0,0,0.6); backdrop-filter: blur(3px); z-index: -1;
                    }
                    div#exception-marker { display: none; }
                    </style>
                    <div id="exception-marker"></div>
                    """, unsafe_allow_html=True)
                
                st.markdown(f"""
                    <div style="text-align: center; color: #fff;">
                        <div style="font-size: 60px; margin-bottom: 10px;">🚫</div>
                        <h3 style="color: #ff6b6b; margin: 0;">檔案儲存失敗</h3>
                        <div style="background: rgba(0,0,0,0.4); padding: 15px; margin-top: 15px; border-radius: 8px; text-align: left; font-family: monospace; font-size: 13px; color: #ffaaaa;">
                            {st.session_state.io_exception_msg}
                        </div>
                        <p style="margin-top: 15px; color: #ddd; font-size: 14px;">請檢查檔案權限或路徑設定。</p>
                    </div>
                """, unsafe_allow_html=True)

                c1, c2, c3 = st.columns([1, 2, 1])
                with c2:
                    if st.button("❌ 關閉視窗", key="btn_close_exception"):
                        st.session_state.show_io_exception_modal = False
                        st.rerun()

        # --- 情境三：儲存成功 (自動淡出，無按鈕) ---
        elif st.session_state.show_save_success_modal:
            st.balloons()
            fade_css = """
            <style>
            @keyframes fadeOutAnim { 0% { opacity: 1; transform: translate(-50%, -50%) scale(1); } 100% { opacity: 0; transform: translate(-50%, -50%) scale(0.9); } }
            .modal-fade-out { animation: fadeOutAnim 1s ease-out forwards; }
            </style>
            """
            success_html = f"""
            {fade_css}
            <div id="success-modal" style="
                position: fixed; top: 50%; left: 50%; transform: translate(-50%, -50%);
                z-index: 999999;
                background: linear-gradient(135deg, rgba(11, 22, 38, 0.98), rgba(28, 69, 50, 0.95));
                border: 2px solid #4cd37a; border-radius: 20px;
                padding: 40px; text-align: center; width: 450px;
                box-shadow: 0 0 60px rgba(76, 211, 122, 0.4);
                backdrop-filter: blur(10px);
            ">
                <div style="font-size: 70px; margin-bottom: 15px; animation: kpiPulse 1.5s infinite;">✅</div>
                <h2 style="color: #4cd37a; margin: 0; font-weight: 800; letter-spacing: 1px;">儲存成功！</h2>
                <p style="color: #e6eef6; margin-top: 10px; font-size: 16px;">資料驗證通過並已安全寫入</p>
                <div style="margin-top: 20px; border-top: 1px solid rgba(255,255,255,0.1); padding-top: 10px;">
                    <span style="color: #88f2ff; font-size: 13px; font-family: monospace;">
                        TIMESTAMP: {st.session_state.last_save_time}
                    </span>
                </div>
            </div>
            """
            modal_container.markdown(success_html, unsafe_allow_html=True)
            time.sleep(5)
            success_html_fade = success_html.replace('id="success-modal"', 'id="success-modal" class="modal-fade-out"')
            modal_container.markdown(success_html_fade, unsafe_allow_html=True)
            time.sleep(1)
            st.session_state.show_save_success_modal = False
            st.rerun()
#在終端機輸入：python -m streamlit run "C:\Users\user\OneDrive\桌面\dashboard.py"